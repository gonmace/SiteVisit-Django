from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Case, IntegerField, Q, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView
from django.views.generic.edit import CreateView, DeleteView, UpdateView

from core.mixins import (
    PortalAccessRequiredMixin, SuperManagerRequiredMixin, CompanyScopedQuerysetMixin,
)
from sites.forms import SiteForm
from sites.models import Site
from users.models import User


class SitesListView(PortalAccessRequiredMixin, CompanyScopedQuerysetMixin, ListView):
    template_name = 'manager/sites_list.html'
    context_object_name = 'sites'
    paginate_by = 50

    def get_queryset(self):
        qs = Site.objects.filter(is_active=True).annotate(
            sin_empresa=Case(
                When(company='', then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).order_by('sin_empresa', 'code')
        qs = self.apply_company_filter(qs, 'company')
        q = self.request.GET.get('q', '').strip()
        if q:
            qs = qs.filter(Q(code__icontains=q) | Q(operator_code__icontains=q) | Q(name__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q', '')
        ctx.update(self.company_context())
        ctx['is_super_manager'] = (
            self.request.user.role == User.Role.SUPER_MANAGER or self.request.user.is_superuser
        )
        if ctx['is_super_manager']:
            for site in ctx['sites']:
                site.edit_form = SiteForm(instance=site)
        return ctx

    def render_to_response(self, context, **response_kwargs):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return render(self.request, 'manager/partials/sites_results.html', context)
        return super().render_to_response(context, **response_kwargs)


class SiteDetailView(PortalAccessRequiredMixin, View):
    def get_queryset(self):
        return Site.objects.all()

    def get(self, request, pk, **kwargs):
        site = get_object_or_404(self.get_queryset(), pk=pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'code': site.code,
                'operator_code': site.operator_code or '',
                'name': site.name,
                'latitude': float(site.latitude) if site.latitude is not None else None,
                'longitude': float(site.longitude) if site.longitude is not None else None,
                'height': site.height,
                'company': site.get_company_display(),
                'is_active': site.is_active,
            })
        return redirect('manager:sites_list')


class SiteCreateView(SuperManagerRequiredMixin, CreateView):
    model = Site
    form_class = SiteForm
    template_name = 'manager/site_form.html'
    success_url = reverse_lazy('manager:sites_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Agregar Sitio'
        ctx['cancel_url'] = reverse_lazy('manager:sites_list')
        return ctx


class SiteUpdateView(SuperManagerRequiredMixin, View):
    def _get_site(self, pk):
        return get_object_or_404(Site, pk=pk)

    def get(self, request, pk, **kwargs):
        site = self._get_site(pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            form = SiteForm(instance=site)
            return render(request, 'manager/site_form_modal.html', {'form': form, 'site': site})
        return redirect('manager:sites_list')

    def post(self, request, pk, **kwargs):
        site = self._get_site(pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            form = SiteForm(request.POST, instance=site)
            if form.is_valid():
                form.save()
                messages.success(request, f'Sitio {site.code} actualizado.')
                return JsonResponse({'ok': True})
            return render(request, 'manager/site_form_modal.html', {'form': form, 'site': site}, status=422)
        return redirect('manager:sites_list')


class SiteDeleteView(SuperManagerRequiredMixin, DeleteView):
    model = Site
    template_name = 'manager/site_confirm_delete.html'
    success_url = reverse_lazy('manager:sites_list')

    def form_valid(self, form):
        code = self.object.code
        self.object.is_active = False
        self.object.save(update_fields=['is_active'])
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'message': f'Sitio {code} eliminado.'})
        messages.success(self.request, f'Sitio {code} eliminado.')
        return redirect(self.success_url)


class SiteSearchJsonView(PortalAccessRequiredMixin, CompanyScopedQuerysetMixin, View):
    """Returns JSON list of sites matching ?q= for autocomplete in visit form."""
    def get(self, request):
        q = request.GET.get('q', '').strip()
        if len(q) < 3:
            return JsonResponse([], safe=False)
        qs = self.apply_company_filter(Site.objects.order_by('code'), 'company')
        qs = qs.filter(Q(code__icontains=q) | Q(operator_code__icontains=q) | Q(name__icontains=q))
        data = [{'id': s.pk, 'code': s.code, 'operator_code': s.operator_code, 'name': s.name,
                 'lat': s.latitude, 'lng': s.longitude} for s in qs[:20]]
        return JsonResponse(data, safe=False)


class SiteImportView(SuperManagerRequiredMixin, View):
    template_name = 'manager/site_import.html'

    _col_spec = [
        ('A', 'Código Sitio',    True),
        ('B', 'Cód. Operador',   False),
        ('C', 'Nombre Sitio',    True),
        ('D', 'Latitud',         True),
        ('E', 'Longitud',        True),
        ('F', 'Altura (m)',      False),
        ('G', 'Comuna',          False),
        ('H', 'Región',          False),
        ('I', 'Empresa (wom/pti)', True),
        ('J', 'Activo (Sí/No)',  False),
    ]

    def get(self, request):
        return render(request, self.template_name, {'col_spec': self._col_spec})

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Selecciona un archivo Excel.')
            return redirect('manager:site_import')

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active

            created = updated = 0
            errors = []
            seen_codes = set()

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    # Columnas: Código | Cód.Op. | Nombre | Lat | Lon | Altura | Comuna | Región | Empresa | Activo
                    code      = str(row[0]).strip() if row[0] else ''
                    op_code   = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    name      = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    latitude  = float(row[3]) if len(row) > 3 and row[3] is not None else None
                    longitude = float(row[4]) if len(row) > 4 and row[4] is not None else None
                    height    = int(row[5]) if len(row) > 5 and row[5] is not None else None
                    comuna    = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    region    = str(row[7]).strip() if len(row) > 7 and row[7] else ''
                    company   = str(row[8]).strip().lower() if len(row) > 8 and row[8] else ''

                    if not code or not name or latitude is None or longitude is None:
                        errors.append(f'Fila {i}: código, nombre, latitud y longitud son requeridos.')
                        continue

                    if code in seen_codes:
                        errors.append(f'Fila {i}: código "{code}" duplicado en el archivo, se omite.')
                        continue
                    seen_codes.add(code)

                    existing = Site.objects.filter(code=code).first()
                    if not existing and not company:
                        errors.append(f'Fila {i}: la empresa es requerida para sitios nuevos.')
                        continue
                    if company and company not in [c[0] for c in User.Company.choices]:
                        errors.append(f'Fila {i}: empresa "{company}" no válida (wom, pti).')
                        continue

                    defaults = {
                        'operator_code': op_code,
                        'name': name,
                        'latitude': latitude,
                        'longitude': longitude,
                        'height': height,
                        'comuna': comuna,
                        'region': region,
                    }
                    if company:
                        defaults['company'] = company

                    _, flag = Site.objects.update_or_create(code=code, defaults=defaults)
                    if flag:
                        created += 1
                    else:
                        updated += 1

                except Exception as e:
                    errors.append(f'Fila {i}: {e}')

            msg = f'Importación completada: {created} creados, {updated} actualizados.'
            if errors:
                for err in errors[:5]:
                    messages.warning(request, err)
            messages.success(request, msg)

        except Exception as e:
            messages.error(request, f'Error al leer el archivo: {e}')

        return redirect('manager:sites_list')


class SiteTemplateView(SuperManagerRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sitios'

        headers = ['Código', 'Cód. Operador', 'Nombre', 'Latitud', 'Longitud', 'Altura (m)', 'Comuna', 'Región', 'Empresa', 'Activo']
        header_fill = PatternFill('solid', fgColor='4A4D4E')
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        col_widths = [16, 16, 40, 14, 14, 14, 20, 28, 10, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_sitios.xlsx"'
        return response


class SiteExportView(SuperManagerRequiredMixin, View):
    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        qs = Site.objects.order_by('company', 'code')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sitios'

        headers = ['Código', 'Cód. Operador', 'Nombre', 'Latitud', 'Longitud', 'Altura (m)', 'Comuna', 'Región', 'Empresa', 'Activo']
        header_fill = PatternFill('solid', fgColor='4A4D4E')
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, site in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=site.code)
            ws.cell(row=row_idx, column=2, value=site.operator_code or '')
            ws.cell(row=row_idx, column=3, value=site.name)
            ws.cell(row=row_idx, column=4, value=float(site.latitude) if site.latitude is not None else None)
            ws.cell(row=row_idx, column=5, value=float(site.longitude) if site.longitude is not None else None)
            ws.cell(row=row_idx, column=6, value=site.height)
            ws.cell(row=row_idx, column=7, value=site.comuna or '')
            ws.cell(row=row_idx, column=8, value=site.region or '')
            ws.cell(row=row_idx, column=9, value=site.company)
            ws.cell(row=row_idx, column=10, value='Sí' if site.is_active else 'No')

        col_widths = [16, 16, 40, 14, 14, 14, 20, 28, 10, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="sitios.xlsx"'
        return response
